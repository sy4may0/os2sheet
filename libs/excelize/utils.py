from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from typing import Union, Dict, Any, Optional


def column_increment(col: str, offset: int = 1) -> str:
    """
    Increment the column letter by the given offset.

    Args:
        col (str): The column letter to increment.
        offset (int): The number of columns to increment.

    Returns:
        str: The incremented column letter.

    Raises:
        ValueError: If the column letter is invalid.
    """
    return get_column_letter(column_index_from_string(col) + offset)


def update_cell_border(cell: Cell, **sides: Dict[str, Optional[Side]]) -> None:
    """
    Update the border of the cell.

    Args:
        cell (Cell): The cell to update.
        **sides (Dict[str, Optional[Side]]): The sides to update.
            Keys can be: 'top', 'bottom', 'left', 'right', 'diagonal',
            'diagonal_direction', 'vertical', 'horizontal'

    Raises:
        ValueError: If the cell is invalid.
    """
    old = cell.border
    top = cell.border.top
    bottom = cell.border.bottom
    left = cell.border.left
    right = cell.border.right

    if sides.get('top') and cell.border.top != sides['top']:
        top = sides.get('top')

    if sides.get('bottom') and cell.border.bottom != sides['bottom']:
        bottom = sides.get('bottom')

    if sides.get('left') and cell.border.left != sides['left']:
        left = sides.get('left')

    if sides.get('right') and cell.border.right != sides['right']:
        right = sides.get('right')

    cell.border = Border(
        top=top, bottom=bottom, left=left, right=right,
        diagonal=old.diagonal,
        diagonal_direction=old.diagonal_direction,
        vertical=old.vertical,
        horizontal=old.horizontal
    )


def __fix_col_index(col: Union[str, int]) -> int:
    """
    Convert column reference to index.

    Args:
        col (Union[str, int]): Column reference (letter or index).

    Returns:
        int: Column index.

    Raises:
        ValueError: If the column reference is invalid.
    """
    if isinstance(col, str):
        return int(column_index_from_string(col))
    return col


def apply_border(
    worksheet: Worksheet,
    start_col: Union[str, int],
    end_col: Union[str, int],
    start_row: int,
    end_row: int,
    color: str = '000000',
    style: str = 'thin',
) -> None:
    """
    Apply a border to a range of cells.

    Args:
        worksheet (Worksheet): The worksheet to apply the border to.
        start_row (int): The start row of the range.
        start_col (Union[str, int]): The start column of the range.
        end_row (int): The end row of the range.
        end_col (Union[str, int]): The end column of the range.
        color (str): The color of the border.
        style (str): The style of the border.

    Raises:
        ValueError: If the range is invalid.
    """
    start_col = __fix_col_index(start_col)
    end_col = __fix_col_index(end_col)

    if start_row > end_row or start_col > end_col:
        raise ValueError(
            "Invalid range: start position must be less than end position")

    border_side = Side(style=style, color=color)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)

            top = border_side if row == start_row else None
            bottom = border_side if row == end_row else None
            left = border_side if col == start_col else None
            right = border_side if col == end_col else None

            update_cell_border(
                cell, top=top, bottom=bottom, left=left, right=right
            )


def apply_fill(
    worksheet: Worksheet,
    start_col: Union[str, int],
    end_col: Union[str, int],
    start_row: int,
    end_row: int,
    color: str = '000000',
) -> None:
    """
    Apply a fill to a range of cells.

    Args:
        worksheet (Worksheet): The worksheet to apply the fill to.
        start_row (int): The start row of the range.
        start_col (Union[str, int]): The start column of the range.
        end_row (int): The end row of the range.
        end_col (Union[str, int]): The end column of the range.
        color (str): The color of the fill.

    Raises:
        ValueError: If the range is invalid.
    """
    start_col = __fix_col_index(start_col)
    end_col = __fix_col_index(end_col)

    if start_row > end_row or start_col > end_col:
        raise ValueError(
            "Invalid range: start position must be less than end position")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(fill_type='solid', fgColor=color)


def apply_font(
    worksheet: Worksheet,
    start_col: Union[str, int],
    end_col: Union[str, int],
    start_row: int,
    end_row: int,
    font: Font,
) -> None:
    """
    Apply a font to a range of cells.

    Args:
        worksheet (Worksheet): The worksheet to apply the font to.
        start_row (int): The start row of the range.
        start_col (Union[str, int]): The start column of the range.
        end_row (int): The end row of the range.
        end_col (Union[str, int]): The end column of the range.
        font (Font): The font to apply.

    Raises:
        ValueError: If the range is invalid.
    """
    start_col = __fix_col_index(start_col)
    end_col = __fix_col_index(end_col)

    if start_row > end_row or start_col > end_col:
        raise ValueError(
            "Invalid range: start position must be less than end position")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = font
