from typing import List, Dict, Any
from datetime import datetime
from libs.excelize.contents.ContentSheet import ContentSheet
from libs.excelize.contents.Collection import Collection
from libs.excelize.utils import column_increment
from importlib.resources import files
from openpyxl.utils import column_index_from_string
import openpyxl


HEADER_SHEET_NAME = "表紙"
REVISION_SHEET_NAME = "改版履歴"
TOC_SHEET_NAME = "目次"
TMPL_CONTENT_SHEET_NAME = "sheet1"

CELLS_HEADER = {
    'client_name': 'G3',
    'contractor_name': 'G4',
    'document_title': 'G6',
    'project_name': 'G9',
    'document_number': 'G12',
    'system_name': 'G15',
    'document_name': 'G18',
}

STYLES_REVISION = {
    'version': 'C',
    'date': 'H',
    'overview': 'N',
    'details': 'X',
    'author': 'BB',
    'row_begin': 8,
}

STYLES_TOC = {
    'h1': 'C',
    'h2': 'D',
    'end': 'T',
    'width': 18,
    'max_row': 36,
    'row_begin': 6,
}

CONTENT_ROW_BEGIN = 5


class ExcelizerException(Exception):
    pass


class Revision:
    def __init__(
        self,
        version: str = "1.0",
        date: str = "",
        overview: str = "",
        details: str = "",
        author: str = "",
    ):
        """
        Initialize a Revision object.

        Args:
            version (str): The version of the revision.
            date (str): The date of the revision.
            overview (str): The overview of the revision.
            details (str): The details of the revision.
            author (str): The author of the revision.
        """
        self.__version = version
        self.__date = date
        self.__overview = overview
        self.__details = details
        self.__author = author

        if not self.date:
            self.__date = datetime.now().strftime("%Y-%m-%d")
        if not self.__overview:
            self.__overview = '新規作成'
        if not self.__author:
            self.__author = 'N/A'

    @property
    def version(self) -> str:
        return self.__version

    @property
    def date(self) -> str:
        return self.__date

    @property
    def overview(self) -> str:
        return self.__overview

    @property
    def details(self) -> str:
        return self.__details

    @property
    def author(self) -> str:
        return self.__author

    def __repr__(self) -> str:
        return (
            f"Revision(version={self.version}, date={self.date}, "
            f"overview={self.overview}, details={self.details}, "
            f"author={self.author})"
        )


class Excelizer:
    def __init__(
        self,
        write_file: str,
        client_name: str = "",
        contractor_name: str = "",
        project_name: str = "",
        document_number: str = "",
        system_name: str = "",
        document_name: str = "",
        document_title: str = "",
        sheets: List[ContentSheet] = [],
        revisions: List[Revision] = [],
    ):
        """
        Initialize an Excelizer object.

        Args:
            write_file (str): The file to write the Excel document to.
            client_name (str): The name of the client.
            contractor_name (str): The name of the contractor.
            project_name (str): The name of the project.
            document_number (str): The number of the document.
            system_name (str): The name of the system.
            document_name (str): The name of the document.
            document_title (str): The title of the document.
            sheets (List[ContentSheet]): The sheets to include in the Excel document.
            revisions (List[Revision]): The revisions to include in the Excel document.
        """
        self.__write_file = write_file
        self.__client_name = client_name
        self.__contractor_name = contractor_name
        self.__system_name = system_name
        self.__document_name = document_name
        self.__document_title = document_title
        self.__project_name = project_name
        self.__document_number = document_number
        self.__sheets = sheets
        self.__revisions = revisions
        self.__toc = {}
        self.__workbook = None

        self.__init_workbook()

        if not self.__revisions:
            self.__init_revision()

    def __init_workbook(self) -> None:
        """
        Initialize the workbook.
        """
        tmpl_xlsx = files(
            'libs.excel_templates'
        ).joinpath('Template.xlsx')
        tmpl_workbook = openpyxl.load_workbook(tmpl_xlsx)
        tmpl_workbook.save(self.__write_file)
        self.__workbook = openpyxl.load_workbook(self.__write_file)

    def __init_revision(self) -> None:
        """
        Initialize the revision.
        """
        self.__revisions.append(
            Revision()
        )

    def fix_sheet_index(self) -> None:
        """
        Fix the index of the sheet.
        """
        i = 1
        for sheet in self.__sheets:
            sheet.set_index(i)
            i += 1

    def build_toc(self) -> None:
        """
        Build the table of contents.
        Reassign the index of the sheet and collection.

        Returns:
            Dict[str, List[str]]: The table of contents.

        Raises:
            ExcelizerException: If sheets have not been set.
        """
        if not self.__sheets:
            raise ExcelizerException("Sheets have not been set.")

        for sheet in self.__sheets:
            self.fix_sheet_index()
            sheet.fix_contents_index()
            toc_h1_key = f"{sheet.index}. {sheet.key}"
            self.__toc[toc_h1_key] = []

            for collection in sheet.items:
                if not isinstance(collection, Collection):
                    continue

                toc_h2_key = (
                    f"{sheet.index}.{collection.index}. "
                    f"{collection.key}"
                )
                self.__toc[toc_h1_key].append(toc_h2_key)

    def add_sheet(self, sheet: ContentSheet) -> None:
        """
        Add a sheet to the Excelizer object.
        """
        self.__sheets.append(sheet)
        self.build_toc()

    def add_revision(self, revision: Revision) -> None:
        """
        Add a revision to the Excelizer object.
        """
        self.__revisions.append(revision)

    def __write_header(self) -> None:
        """
        Write the header of the Excel document.
        """
        if not self.__workbook:
            raise ExcelizerException("Workbook has not been set.")

        sheet = self.__workbook[HEADER_SHEET_NAME]

        sheet[CELLS_HEADER['client_name']] = self.__client_name
        sheet[CELLS_HEADER['contractor_name']] = self.__contractor_name
        sheet[CELLS_HEADER['document_title']] = self.__document_title
        sheet[CELLS_HEADER['project_name']] = self.__project_name
        sheet[CELLS_HEADER['document_number']] = self.__document_number
        sheet[CELLS_HEADER['system_name']] = self.__system_name
        sheet[CELLS_HEADER['document_name']] = self.__document_name

    def __write_revision(self) -> None:
        """
        Write the revision of the Excel document.
        """
        if not self.__workbook:
            raise ExcelizerException("Workbook has not been set.")

        if not self.__revisions:
            raise ExcelizerException("Revisions have not been set.")

        revision_sheet = self.__workbook[REVISION_SHEET_NAME]
        row = STYLES_REVISION['row_begin']
        for revision in self.__revisions:
            revision_sheet[f"{STYLES_REVISION['version']}{row}"] = float(
                revision.version)
            revision_sheet[f"{STYLES_REVISION['date']}{row}"] = revision.date
            revision_sheet[f"{STYLES_REVISION['overview']}{row}"] = revision.overview
            revision_sheet[f"{STYLES_REVISION['details']}{row}"] = revision.details
            revision_sheet[f"{STYLES_REVISION['author']}{row}"] = revision.author
            row += 1

    def __write_toc(self) -> None:
        """
        Write the table of contents of the Excel document.
        """
        if not self.__workbook:
            raise ExcelizerException("Workbook has not been set.")

        if not self.__toc:
            raise ExcelizerException("Table of contents has not been set.")

        toc_sheet = self.__workbook[TOC_SHEET_NAME]
        row = STYLES_TOC['row_begin']
        h1_col = STYLES_TOC['h1']
        h2_col = STYLES_TOC['h2']
        end_col = STYLES_TOC['end']
        for toc_h1_key, toc_h2_keys in self.__toc.items():
            if row + len(toc_h2_keys) > STYLES_TOC['max_row']:
                h1_col = column_increment(h1_col, STYLES_TOC['width'])
                h2_col = column_increment(h2_col, STYLES_TOC['width'])
                end_col = column_increment(end_col, STYLES_TOC['width'])
                row = STYLES_TOC['row_begin']

            toc_sheet[f"{h1_col}{row}"] = toc_h1_key
            toc_sheet.merge_cells(
                start_row=row, start_column=column_index_from_string(h1_col),
                end_row=row, end_column=column_index_from_string(end_col)
            )
            row += 1
            for toc_h2_key in toc_h2_keys:
                toc_sheet[f"{h2_col}{row}"] = toc_h2_key
                toc_sheet.merge_cells(
                    start_row=row, start_column=column_index_from_string(
                        h2_col),
                    end_row=row, end_column=column_index_from_string(end_col)
                )
                row += 1

    def __write_sheet(self, sheet: ContentSheet) -> None:
        """
        Write the sheet of the Excel document.
        """
        if not self.__workbook:
            raise ExcelizerException("Workbook has not been set.")
        if not self.__sheets:
            raise ExcelizerException("Sheets have not been set.")

        base_sheet = self.__workbook[TMPL_CONTENT_SHEET_NAME]
        new_sheet = self.__workbook.copy_worksheet(base_sheet)
        new_sheet.title = sheet.sheetname

        sheet.write(self.__workbook[sheet.sheetname], CONTENT_ROW_BEGIN)

    def write(self) -> None:
        """
        Write the Excel document.
        """
        self.__write_header()
        self.__write_revision()
        self.__write_toc()

        for sheet in self.__sheets:
            self.__write_sheet(sheet)

    def save_and_close(self) -> None:
        """
        Save and close the Excel document.
        """
        if not self.__workbook:
            raise ExcelizerException("Workbook has not been set.")

        self.__workbook.remove(self.__workbook[TMPL_CONTENT_SHEET_NAME])
        self.__workbook.save(self.__write_file)
        self.__workbook.close()

    def get_dict(self) -> List[Dict[str, Any]]:
        """
        Get the dictionary of the Excel document.
        """
        return [d.get_dict() for d in self.__sheets]
